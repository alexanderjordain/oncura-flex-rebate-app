"""Tests for the monthly FLEX audit workbook builder."""
import io

from openpyxl import load_workbook

from core import monthly_audit as ma


def _payments():
    return [
        # June cycle, GreatAmerica: finance payment + its credit memo (paired -> no flag)
        {"kind": "flex", "company": "GreatAmerica", "qb_customer": "Clinic A",
         "payment_date": "2026-06-15", "amount": 900.0, "contract": "022-1", "fingerprint": "aaaa1111"},
        {"kind": "credit_memo", "company": "INTERNAL", "qb_customer": "Clinic A",
         "payment_date": "06/30/2026", "amount": 1000.0, "contract": "", "fingerprint": "bbbb2222"},
        # NewLane received in July but covers May -> attributed to June cycle
        {"kind": "flex", "company": "NewLane", "qb_customer": "Clinic B",
         "payment_date": "2026-07-07", "applies_to": "2026-05", "amount": 500.0,
         "contract": "40010", "fingerprint": "cccc3333"},
        # OnePlace in May -> NOT June
        {"kind": "flex", "company": "OnePlace", "qb_customer": "Clinic A",
         "payment_date": "2026-05-20", "amount": 800.0, "contract": "04001", "fingerprint": "dddd4444"},
        # A negative adjustment in June
        {"kind": "flex", "company": "GreatAmerica", "qb_customer": "Clinic C",
         "payment_date": "2026-06-02", "amount": -50.0, "contract": "022-9", "fingerprint": "eeee5555"},
    ]


def test_ym_handles_both_formats():
    assert ma._ym("2026-06-15") == (2026, 6)
    assert ma._ym("06/30/2026") == (2026, 6)
    assert ma._ym("") is None
    assert ma._ym("garbage") is None


def test_row_ym_newlane_uses_coverage():
    # coverage 2026-05 -> attributed to 2026-06 despite a July payment_date
    row = {"kind": "flex", "company": "NewLane", "payment_date": "2026-07-07",
           "applies_to": "2026-05"}
    assert ma._row_ym(row) == (2026, 6)


def test_categorize_scopes_to_month():
    rows = ma.categorize(_payments(), 2026, 6)
    # GA flex + credit memo + NewLane flex + negative GA = 4; OnePlace May excluded
    assert len(rows) == 4
    assert all(r.get("payment_date") != "2026-05-20" for r in rows)


def test_summarize_flags_and_totals():
    roster = [{"qb_name": "Clinic A", "finance_company": "GreatAmerica",
               "monthly_credit": 1000.0, "quarterly_threshold": 6000.0}]
    rows = ma.categorize(_payments(), 2026, 6)
    summary, review, totals = ma.summarize(rows, roster)

    by = {r["clinic"]: r for r in summary}
    # Clinic A: paired payment + credit memo, on roster -> no flags
    assert by["Clinic A"]["flex"] == 900.0 and by["Clinic A"]["credit_memo"] == 1000.0
    assert by["Clinic A"]["review"] == ""
    assert by["Clinic A"]["monthly_credit"] == 1000.0
    # Clinic B: NewLane payment, no credit memo, off roster -> two flags
    assert "NO credit memo" in by["Clinic B"]["review"]
    assert "roster" in by["Clinic B"]["review"]
    # Clinic C: negative amount
    assert "Negative" in by["Clinic C"]["review"]

    assert totals["clinics"] == 3
    assert totals["finance_total"] == round(900.0 + 500.0 - 50.0, 2)
    assert totals["credit_total"] == 1000.0
    review_clinics = {r["clinic"] for r in review}
    assert review_clinics == {"Clinic B", "Clinic C"}


def test_scan_only_customer_excluded_from_summary():
    pays = [
        {"kind": "flex", "company": "GreatAmerica", "qb_customer": "Clinic A",
         "payment_date": "2026-06-15", "amount": 900.0, "fingerprint": "f1"},
        # scan-only customer, not on the roster -> kept out of the summary
        {"kind": "scan", "company": "NewLane", "qb_customer": "Scan Only Co",
         "payment_date": "2026-06-10", "amount": 300.0, "fingerprint": "f2"},
    ]
    roster = [{"qb_name": "Clinic A", "finance_company": "GreatAmerica", "monthly_credit": 1000.0}]
    summary, review, totals = ma.summarize(ma.categorize(pays, 2026, 6), roster)
    names = {r["clinic"] for r in summary}
    assert "Clinic A" in names
    assert "Scan Only Co" not in names          # excluded from per-clinic summary
    assert totals["scan_total"] == 300.0        # but still counted in the scan total
    assert totals["clinics"] == 1


def test_build_workbook_has_four_tabs():
    roster = [{"qb_name": "Clinic A", "finance_company": "GreatAmerica",
               "monthly_credit": 1000.0}]
    xbytes, totals = ma.build_workbook(2026, 6, roster, payments=_payments(),
                                       generated="2026-07-13")
    assert isinstance(xbytes, bytes) and len(xbytes) > 0
    wb = load_workbook(io.BytesIO(xbytes))
    assert wb.sheetnames == ["Read me", "Summary by clinic", "Ledger detail", "Review"]
    # Ledger detail should have a header + 4 entry rows
    ws = wb["Ledger detail"]
    assert ws["A3"].value == "Clinic"
    assert totals["clinics"] == 3


def test_build_workbook_empty_month():
    xbytes, totals = ma.build_workbook(2026, 1, [], payments=_payments())
    assert len(xbytes) > 0 and totals["clinics"] == 0
    wb = load_workbook(io.BytesIO(xbytes))
    assert "No review items" in (wb["Review"]["A3"].value or "")
