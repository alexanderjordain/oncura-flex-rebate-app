"""One-off / re-runnable: build data/rebate_master.json from the 'Rebate Names' tab of
the existing Rebate Accounts workbook.

Program type + rates are inferred from the Finance Company column:
  - 'Self-Financed'                  -> self_funded, ultrasound 5%
  - 'OnePlace Capital' / 'NewLane*'  -> finance,     ultrasound 10%
  - rads 4% across the board (self_funded rads is an OPEN QUESTION; flagged per clinic)

Usage:
  python scripts/build_rebate_master.py "<path to Rebate Accounts Copy.xlsx>"
"""
import io
import json
import os
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from openpyxl import load_workbook

DEFAULT_WB = r"C:\Users\AlexanderJordain\OneDrive - Oncura Partners\Rebate Accounts Copy.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "rebate_master.json")

RATE_US_SELF = 0.05
RATE_US_FINANCE = 0.10
RATE_RADS_FINANCE = 0.04
RATE_RADS_SELF = 0.02  # self-funded rads = half of finance, per OPD feed (decision 2026-05-26)


def normalize_finance_company(raw: str) -> str:
    v = (raw or "").strip()
    low = v.lower()
    if low.startswith("self"):
        return "Self-Financed"
    if low.startswith("oneplace"):
        return "OnePlace Capital"
    if low.startswith("newlane"):
        return "NewLane Financed"
    return v


def program_for(finance_company: str) -> str:
    return "self_funded" if finance_company == "Self-Financed" else "finance"


def build(wb_path: str):
    wb = load_workbook(wb_path, data_only=True)
    ws = wb["Rebate Names"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    clinics = []
    for r in rows:
        legal, clinic, finance_raw = r[0], r[1], r[2]
        if not (legal or clinic):
            continue
        fc = normalize_finance_company(finance_raw)
        program = program_for(fc)
        clinics.append(
            {
                "legal_name": (str(legal).strip() if legal else None),
                "clinic_name": (str(clinic).strip() if clinic else None),
                "finance_company": fc,
                "program_type": program,
                "rate_ultrasound": RATE_US_SELF if program == "self_funded" else RATE_US_FINANCE,
                # rads: finance 4%, self-funded 2% (half of finance, per OPD feed RadCash).
                # Decision 2026-05-26; adjustable per-clinic in the Rebate Master UI.
                "rate_rads": RATE_RADS_SELF if program == "self_funded" else RATE_RADS_FINANCE,
                "rads_rate_confirmed": True,
                "rads_rate_note": "",
                "active": True,
                "notes": "",
            }
        )

    clinics.sort(key=lambda c: (c["clinic_name"] or c["legal_name"] or "").lower())
    payload = {
        "version": 1,
        "source": "Rebate Accounts Copy.xlsx :: 'Rebate Names' tab",
        "rate_defaults": {
            "ultrasound_self_funded": RATE_US_SELF,
            "ultrasound_finance": RATE_US_FINANCE,
            "rads_self_funded": RATE_RADS_SELF,
            "rads_finance": RATE_RADS_FINANCE,
        },
        "clinics": clinics,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(os.path.normpath(OUT), "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    by_prog = {}
    for c in clinics:
        by_prog[c["program_type"]] = by_prog.get(c["program_type"], 0) + 1
    by_fc = {}
    for c in clinics:
        by_fc[c["finance_company"]] = by_fc.get(c["finance_company"], 0) + 1
    print(f"Wrote {len(clinics)} clinics to {os.path.normpath(OUT)}")
    print("By program:", by_prog)
    print("By finance company:", by_fc)


if __name__ == "__main__":
    wb_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WB
    build(wb_path)
