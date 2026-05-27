"""Build data/name_map.json: finance-company / legal customer name -> QBO customer (payee).

Finance remittances list legal entity names (e.g. "AKG Vet, Inc.") but QBO/imports need the
QB payee name (e.g. "Pinnacle Animal Hospital"). Source: 'ScanPackage' tab (Customer Name ->
QB Payee) plus FlexMaster (Clinic Name / QBName) for broader coverage.

Usage: python scripts/build_name_map.py "<path to Flex Master List.xlsx>"
"""
import io
import json
import os
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from openpyxl import load_workbook

DEFAULT = (
    r"C:\Users\AlexanderJordain\Oncura Partners\Oncura Partners Accounting Library - Documents"
    r"\Pass Thru Payments\Flex Master List.xlsx"
)
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "name_map.json")


def build(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    mapping = {}

    ws = wb["ScanPackage"]
    for legal, qb in ws.iter_rows(min_row=2, values_only=True):
        if legal and qb:
            mapping[str(legal).strip()] = str(qb).strip()

    # FlexMaster adds Clinic Name -> QBName where not already mapped
    fm = wb["FlexMaster"]
    rows = list(fm.iter_rows(min_row=2, values_only=True))
    for r in rows:
        clinic, qbname = r[0], r[1]
        if clinic and qbname:
            mapping.setdefault(str(clinic).strip(), str(qbname).strip())
    wb.close()

    payload = {
        "version": 1,
        "source": "Flex Master List.xlsx :: ScanPackage (+ FlexMaster fallback)",
        "note": "finance/legal customer name -> QBO payee. Used to translate remittance names.",
        "map": mapping,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(os.path.normpath(OUT), "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    print(f"Wrote {len(mapping)} name mappings to {os.path.normpath(OUT)}")


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else DEFAULT)
