"""Build data/flex_master.json from 'Flex Master List.xlsx' :: FlexMaster tab.

Per-clinic FLEX config the app needs for credit memos + overage/unused recapture:
  monthly_credit  = FlexCredit   (the monthly credit-memo amount)
  monthly_finance_payment = FlexPayment (what the finance company pays monthly)
  monthly_threshold = FlexTotal  (= FlexPayment + FlexCredit)
  quarterly_threshold = FlexQuarter (= FlexTotal * 3; the overage/unused comparison threshold)
  calendar_spread = FlexQuarterSpread (e.g. "March-April-May")

Usage: python scripts/build_flex_master.py "<path to Flex Master List.xlsx>"
"""
import io
import json
import os
import sys
import datetime as dt

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from openpyxl import load_workbook

DEFAULT = (
    r"C:\Users\AlexanderJordain\Oncura Partners\Oncura Partners Accounting Library - Documents"
    r"\Pass Thru Payments\Flex Master List.xlsx"
)
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "flex_master.json")

# Fixed column positions in the FlexMaster header (two 'Notes' columns make name-lookup unsafe)
COL = {
    "clinic_name": 0, "qb_name": 1, "flex_plan": 9, "clinic_id": 12,
    "ema": 13, "support_ema": 14, "hardware_ema": 15,
    "finance": 16, "flex_payment": 17, "flex_credit": 18, "flex_total": 19,
    "flex_quarter": 20, "ga_contract": 22, "opc_contract": 24, "nl_contract": 25,
    "calendar_spread": 26,
}


def norm_finance(raw):
    v = (str(raw).strip() if raw else "")
    low = v.lower()
    if low.startswith("self"):
        return "SelfFinanced"
    if low.startswith("oneplace") or low.startswith("one place"):
        return "OnePlace"
    if low.startswith("newlane") or low.startswith("new lane"):
        return "NewLane"
    if low.startswith("great"):
        return "GreatAmerica"
    return v


def iso(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    return None


def num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def build(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["FlexMaster"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    clinics = []
    warnings = []
    for r in rows:
        if not r or not r[COL["clinic_name"]]:
            continue
        name = str(r[COL["clinic_name"]]).strip()
        payment = num(r[COL["flex_payment"]])
        credit = num(r[COL["flex_credit"]])
        total = num(r[COL["flex_total"]])
        quarter = num(r[COL["flex_quarter"]])
        rec = {
            "clinic_name": name,
            "qb_name": (str(r[COL["qb_name"]]).strip() if r[COL["qb_name"]] else name),
            "clinic_id": (str(r[COL["clinic_id"]]).strip() if r[COL["clinic_id"]] else None),
            "finance_company": norm_finance(r[COL["finance"]]),
            "monthly_credit": credit,
            "monthly_finance_payment": payment,
            "monthly_threshold": total,
            "quarterly_threshold": quarter,
            "calendar_spread": (str(r[COL["calendar_spread"]]).strip() if r[COL["calendar_spread"]] else None),
            "contract_greatamerica": (str(r[COL["ga_contract"]]).strip() if r[COL["ga_contract"]] else None),
            "contract_oneplace": (str(r[COL["opc_contract"]]).strip() if r[COL["opc_contract"]] else None),
            "contract_newlane": (str(r[COL["nl_contract"]]).strip() if r[COL["nl_contract"]] else None),
            "ema_end": iso(r[COL["ema"]]),
            "support_ema_end": iso(r[COL["support_ema"]]),
            "hardware_ema_end": iso(r[COL["hardware_ema"]]),
            "active": (str(r[COL["flex_plan"]]).strip().lower() == "yes"),
            "notes": "",
        }
        # Integrity check: payment + credit should equal threshold
        if payment is not None and credit is not None and total is not None:
            if abs((payment + credit) - total) > 0.5:
                warnings.append(f"{name}: payment+credit ({payment+credit:.2f}) != threshold ({total:.2f})")
        clinics.append(rec)

    clinics.sort(key=lambda c: c["clinic_name"].lower())
    payload = {
        "version": 1,
        "source": "Flex Master List.xlsx :: FlexMaster",
        "note": "monthly_threshold = finance_payment + credit; quarterly_threshold = monthly_threshold * 3",
        "clinics": clinics,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(os.path.normpath(OUT), "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    by_fc = {}
    for c in clinics:
        by_fc[c["finance_company"]] = by_fc.get(c["finance_company"], 0) + 1
    print(f"Wrote {len(clinics)} flex clinics to {os.path.normpath(OUT)}")
    print("By finance company:", by_fc)
    if warnings:
        print(f"Integrity warnings ({len(warnings)}):")
        for w in warnings[:15]:
            print("  ", w)


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else DEFAULT)
