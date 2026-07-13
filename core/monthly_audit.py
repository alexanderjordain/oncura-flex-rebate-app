"""Monthly FLEX audit workbook — an accountant-ready view of one month's
FLEX activity and the QuickBooks entries it maps to, per clinic.

Nothing here writes to QBO or the ledger — it reads the processed-payments
ledger (the durable record of everything Stages 1-3 emitted) and renders a
multi-tab .xlsx an accountant can open and reconcile without knowing the app.

What a month contains, per clinic (see docs/FLEX_PROGRAM_EXPLAINED.md):
  - Finance-company payment  -> QBO: an UNAPPLIED Receive Payment (a credit
    balance), deposited to Undeposited Funds. Intentionally unapplied (Cash SOP-9).
  - Monthly credit memo       -> QBO: a Credit Memo, item Flex-credits, class
    03-Telemedicine. One credit memo per finance payment ("one in, one out").
  - Scan-package payment      -> QBO: a Receive Payment applied to a scan invoice.
  - Unused-Flex-Credits invoice (quarter-end only) -> QBO: an internal invoice,
    NOT mailed, that recognizes the leftover credit as revenue.
  - Overage bill (quarter-end only) -> billed outside OPD (authorize.net / statement).

Dates: finance/scan payments are grouped into the cycle month they FUND
(NewLane is attributed to its coverage month + 1; everyone else by the received
date). Internal entries (credit memos, unused/overage invoices) are grouped by
their QBO posting date.
"""
from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import ledger

MONEY_FMT = '"$"#,##0.00'

# ledger kind -> (human entry name, what it is in QuickBooks)
ENTRY_META = {
    "flex":           ("Finance-company payment",
                       "Receive Payment — unapplied credit (Undeposited Funds)"),
    "credit_memo":    ("Monthly credit memo",
                       "Credit Memo — item Flex-credits, class 03-Telemedicine"),
    "scan":           ("Scan-package payment",
                       "Receive Payment — applied to scan invoice"),
    "unused_invoice": ("Unused-Flex-Credits invoice (quarter-end)",
                       "Invoice — internal, not mailed (recognizes leftover credit)"),
    "direct_overage": ("Overage bill (quarter-end)",
                       "Billed outside OPD — authorize.net / statement"),
}
# Kinds attributed to the cycle they fund (coverage-aware); everything else is
# grouped by its literal QBO posting date.
_CYCLE_KINDS = {"flex", "scan"}


def _ym(date_str) -> tuple[int, int] | None:
    """Parse a ledger payment_date to (year, month). Tolerates the two formats
    the ledger stores: ISO 'YYYY-MM-DD' (finance rows, direct_overage) and US
    'MM/DD/YYYY' (credit memos, unused invoices). None on junk."""
    s = str(date_str or "").strip()
    if not s:
        return None
    if "-" in s[:10]:                       # ISO YYYY-MM-DD
        parts = s[:10].split("-")
        try:
            if len(parts[0]) == 4:
                return int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            return None
    if "/" in s:                            # US MM/DD/YYYY
        parts = s.split("/")
        try:
            if len(parts) == 3:
                return int(parts[2]), int(parts[0])
        except (ValueError, IndexError):
            return None
    return None


def _row_ym(p: dict) -> tuple[int, int] | None:
    """The (year, month) an entry belongs to. Finance/scan rows use the
    coverage-aware attribution month (so a NewLane remittance lands in the
    cycle it's for); internal entries use their posting date."""
    if p.get("kind") in _CYCLE_KINDS:
        ym = ledger._attribution_ym(p)      # handles NewLane coverage + ISO dates
        if ym:
            return ym
    return _ym(p.get("payment_date"))


def _norm(name) -> str:
    return " ".join(str(name or "").casefold().split())


def categorize(payments: list[dict], year: int, month: int) -> list[dict]:
    """All ledger rows that belong to (year, month), newest-format-tolerant."""
    return [p for p in payments if _row_ym(p) == (year, month)]


def _clinic_index(flex_clinics: list[dict]) -> dict:
    """{normalized qb_name -> clinic dict} for enrichment + roster checks."""
    idx = {}
    for c in flex_clinics or []:
        key = _norm(c.get("qb_name") or c.get("clinic_name"))
        if key:
            idx[key] = c
    return idx


def summarize(rows: list[dict], flex_clinics: list[dict]):
    """Roll month rows up per clinic + flag review items.

    Returns (summary_rows, review_rows, totals):
      summary_rows: one dict per clinic (sorted by name) with per-kind $ + counts,
        the expected monthly credit from the roster, and finance company.
      review_rows:  clinics an accountant should eyeball (payment without a credit
        memo, credit memo without a payment, negative amounts, off-roster).
      totals:       {clinics, finance_total, finance_count, credit_total,
                     credit_count, scan_total, unused_total, overage_total}.
    """
    idx = _clinic_index(flex_clinics)
    by_clinic: dict[str, dict] = {}
    for p in rows:
        name = (p.get("qb_customer") or "").strip() or f"(contract {p.get('contract','?')})"
        amt = round(float(p.get("amount") or 0.0), 2)
        kind = p.get("kind", "?")
        rec = by_clinic.setdefault(name, {
            "clinic": name, "company": p.get("company", ""),
            "flex": 0.0, "flex_n": 0, "credit_memo": 0.0, "credit_memo_n": 0,
            "scan": 0.0, "scan_n": 0, "unused_invoice": 0.0, "direct_overage": 0.0,
            "min_amount": 0.0,
        })
        if kind in ("flex", "scan"):
            rec["company"] = rec["company"] or p.get("company", "")
        if kind == "flex":
            rec["flex"] += amt; rec["flex_n"] += 1
        elif kind == "credit_memo":
            rec["credit_memo"] += amt; rec["credit_memo_n"] += 1
        elif kind == "scan":
            rec["scan"] += amt; rec["scan_n"] += 1
        elif kind == "unused_invoice":
            rec["unused_invoice"] += amt
        elif kind == "direct_overage":
            rec["direct_overage"] += amt
        rec["min_amount"] = min(rec["min_amount"], amt)

    # Totals span ALL rows — scan-only customers still count toward the scan
    # total even though they're kept out of the per-clinic FLEX summary below.
    totals = {"clinics": 0, "finance_total": 0.0, "finance_count": 0,
              "credit_total": 0.0, "credit_count": 0, "scan_total": 0.0,
              "unused_total": 0.0, "overage_total": 0.0}
    for rec in by_clinic.values():
        totals["finance_total"] += rec["flex"]; totals["finance_count"] += rec["flex_n"]
        totals["credit_total"] += rec["credit_memo"]; totals["credit_count"] += rec["credit_memo_n"]
        totals["scan_total"] += rec["scan"]
        totals["unused_total"] += rec["unused_invoice"]
        totals["overage_total"] += rec["direct_overage"]

    summary_rows, review_rows = [], []
    for name in sorted(by_clinic, key=str.lower):
        rec = by_clinic[name]
        clinic = idx.get(_norm(name))
        # A FLEX audit is about FLEX activity. A customer with only a scan-package
        # payment and no place on the FLEX roster is a scan client, not a FLEX
        # clinic — keep them out of the per-clinic summary (their scan payment is
        # still in Ledger detail and the scan total) so it doesn't read as noise.
        has_flex_side = bool(
            rec["flex"] or rec["credit_memo"] or rec["unused_invoice"]
            or rec["direct_overage"] or rec["min_amount"] < 0
        )
        if clinic is None and not has_flex_side:
            continue
        rec["monthly_credit"] = round(float(clinic.get("monthly_credit") or 0.0), 2) if clinic else None
        rec["quarterly_threshold"] = round(float(clinic.get("quarterly_threshold") or 0.0), 2) if clinic else None
        if clinic and clinic.get("finance_company"):
            rec["company"] = clinic["finance_company"]
        # Review flags
        flags = []
        if rec["flex"] and not rec["credit_memo"]:
            flags.append("Finance payment but NO credit memo")
        if rec["credit_memo"] and not rec["flex"]:
            flags.append("Credit memo but NO finance payment")
        if rec["min_amount"] < 0:
            flags.append("Negative amount — likely manual adjustment")
        if clinic is None:
            flags.append("FLEX activity but not on the roster (flex_master)")
        rec["review"] = "; ".join(flags)
        if flags:
            review_rows.append(rec)
        summary_rows.append(rec)

    totals["clinics"] = len(summary_rows)
    for k in ("finance_total", "credit_total", "scan_total", "unused_total", "overage_total"):
        totals[k] = round(totals[k], 2)
    return summary_rows, review_rows, totals


def _detail_rows(rows: list[dict]) -> list[dict]:
    """One flat, sorted row per ledger entry for the audit-trail tab."""
    out = []
    for p in rows:
        kind = p.get("kind", "?")
        entry, qbo = ENTRY_META.get(kind, (kind, ""))
        out.append({
            "clinic": (p.get("qb_customer") or "").strip() or f"(contract {p.get('contract','?')})",
            "entry": entry,
            "qbo": qbo,
            "company": p.get("company", ""),
            "date": str(p.get("payment_date", ""))[:10],
            "amount": round(float(p.get("amount") or 0.0), 2),
            "contract": str(p.get("contract", "")),
            "ledger_id": str(p.get("fingerprint", ""))[:12],
        })
    out.sort(key=lambda r: (r["clinic"].lower(), r["entry"], r["date"]))
    return out


# ── Workbook rendering ──────────────────────────────────────────────────────

_TITLE = Font(bold=True, size=14)
_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="D9DDE3")
_BOX = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_table(ws, headers, rows, start_row, money_cols):
    """Write a header band + rows. `rows` is a list of tuples matching headers.
    money_cols: set of 0-based column indices to format as currency. Returns
    the row after the last data row."""
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = _HEAD; c.fill = _HEAD_FILL; c.alignment = _CTR; c.border = _BOX
    r = start_row + 1
    for tup in rows:
        for j, val in enumerate(tup, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = _BOX
            if (j - 1) in money_cols:
                c.number_format = MONEY_FMT
        r += 1
    return r


def build_workbook(year: int, month: int, flex_clinics: list[dict],
                   payments: list[dict] | None = None,
                   generated: str | None = None):
    """Build the monthly audit .xlsx. Returns (bytes, totals).

    payments: optional ledger rows (for tests); defaults to the live ledger.
    generated: optional 'as of' stamp for the Read-me tab.
    """
    if payments is None:
        data, _ = ledger.load()
        payments = data.get("payments", [])
    rows = categorize(payments, year, month)
    summary_rows, review_rows, totals = summarize(rows, flex_clinics)
    detail = _detail_rows(rows)
    period = dt.date(year, month, 1).strftime("%B %Y")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Tab 1: Read me ────────────────────────────────────────────────────
    ws = wb.create_sheet("Read me")
    ws.cell(row=1, column=1, value=f"FLEX Monthly Audit — {period}").font = _TITLE
    lines = [
        "",
        f"Clinics with activity: {totals['clinics']}",
        f"Finance-company payments: {totals['finance_count']}  (${totals['finance_total']:,.2f})",
        f"Monthly credit memos: {totals['credit_count']}  (${totals['credit_total']:,.2f})",
        f"Scan-package payments: ${totals['scan_total']:,.2f}",
        f"Unused-Flex-Credits invoices (quarter-end): ${totals['unused_total']:,.2f}",
        f"Overage bills (quarter-end): ${totals['overage_total']:,.2f}",
        "",
        "HOW TO READ THIS WORKBOOK",
        "Each FLEX clinic gets TWO entries per month in QuickBooks, not one:",
        "  1. Finance-company payment — cash wired to Oncura by OnePlace / GreatAmerica /",
        "     NewLane, recorded as an UNAPPLIED Receive Payment (a credit balance). It is",
        "     intentionally left unapplied (Cash SOP-9); it is applied at quarter-end.",
        "  2. Monthly credit memo — a Credit Memo, item 'Flex-credits', class",
        "     '03-Telemedicine'. One credit memo is issued per finance payment.",
        "",
        "Together they fund the clinic's quarterly entitlement. At quarter-end the",
        "account is reconciled and the leftover determines the outcome:",
        "  - Leftover credit  -> an internal 'Unused-Flex-Credits' invoice (not mailed)",
        "    that recognizes the credit as revenue.",
        "  - Clinic owes       -> an overage, billed outside OPD.",
        "",
        "TABS",
        "  - Summary by clinic: one row per clinic — finance payment, credit memo, scan,",
        "    unused, overage, plus the expected monthly credit from the roster.",
        "  - Ledger detail: every individual entry, with the QuickBooks entry it maps to.",
        "  - Review: clinics worth a second look (payment without a matching credit memo,",
        "    negative amounts, or names not on the FLEX roster).",
        "",
        "DATES",
        "  Finance and scan payments are grouped into the cycle month they fund (NewLane",
        "  is attributed to its coverage month + 1; all others by the received date).",
        "  Credit memos and quarter-end invoices are grouped by their QuickBooks date.",
        "",
        f"Source: processed-payments ledger. Generated {generated or dt.date.today().isoformat()}.",
        "This workbook is read-only; it does not post anything to QuickBooks.",
    ]
    for i, text in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=text)
    ws.column_dimensions["A"].width = 92

    # ── Tab 2: Summary by clinic ──────────────────────────────────────────
    ws = wb.create_sheet("Summary by clinic")
    ws.merge_cells("A1:K1")
    t = ws.cell(row=1, column=1, value=f"FLEX activity by clinic — {period}")
    t.font = _TITLE; t.alignment = Alignment(horizontal="left")
    headers = ["Clinic", "Finance co", "Finance payment", "# pmts",
               "Credit memo", "# CMs", "Expected monthly credit",
               "Scan", "Unused invoice", "Overage", "Review"]
    money_cols = {2, 4, 6, 7, 8, 9}
    data_rows = [
        (r["clinic"], r["company"], r["flex"], r["flex_n"],
         r["credit_memo"], r["credit_memo_n"],
         r["monthly_credit"] if r["monthly_credit"] is not None else "",
         r["scan"], r["unused_invoice"], r["direct_overage"], r["review"])
        for r in summary_rows
    ]
    end = _write_table(ws, headers, data_rows, start_row=3, money_cols=money_cols)
    # Totals row
    tr = end
    ws.cell(row=tr, column=1, value="Total").font = _BOLD
    for col, val in [(3, totals["finance_total"]), (4, totals["finance_count"]),
                     (5, totals["credit_total"]), (6, totals["credit_count"]),
                     (8, totals["scan_total"]), (9, totals["unused_total"]),
                     (10, totals["overage_total"])]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font = _BOLD; c.border = _BOX
        if col in (3, 5, 8, 9, 10):
            c.number_format = MONEY_FMT
    widths = [34, 13, 16, 8, 14, 8, 16, 12, 14, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # ── Tab 3: Ledger detail ──────────────────────────────────────────────
    ws = wb.create_sheet("Ledger detail")
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value=f"Audit trail — every entry, {period}")
    t.font = _TITLE; t.alignment = Alignment(horizontal="left")
    dheaders = ["Clinic", "Entry", "QuickBooks entry", "Finance co",
                "Date", "Amount", "Contract", "Ledger ID"]
    drows = [(r["clinic"], r["entry"], r["qbo"], r["company"], r["date"],
              r["amount"], r["contract"], r["ledger_id"]) for r in detail]
    _write_table(ws, dheaders, drows, start_row=3, money_cols={5})
    for i, w in enumerate([34, 30, 52, 13, 12, 13, 20, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # ── Tab 4: Review ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Review")
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value=f"Items to review — {period}")
    t.font = _TITLE; t.alignment = Alignment(horizontal="left")
    if review_rows:
        rheaders = ["Clinic", "Finance co", "Finance payment", "Credit memo", "Flag"]
        rrows = [(r["clinic"], r["company"], r["flex"], r["credit_memo"], r["review"])
                 for r in review_rows]
        _write_table(ws, rheaders, rrows, start_row=3, money_cols={2, 3})
        for i, w in enumerate([34, 13, 16, 14, 48], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        ws.cell(row=3, column=1, value="No review items — every clinic's finance "
                "payment has a matching credit memo, and nothing looks off.")
        ws.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), totals
