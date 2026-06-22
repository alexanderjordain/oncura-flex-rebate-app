"""Multi-tab rebate report — one xlsx per cycle, one tab per finance bucket.

Layout (per tab) matches the existing Rebate Accounts Copy.xlsx workbook:

    [row 2]  {Long Period} - {Bucket Pretty Name} Rebates    (merged, bold)
    [row 4]  Clinic/Hospital Name | <Month 1> | <Month 2> | ... | Amount   (bold)
    [row 6+] <Clinic>            |   $X.XX   |   $Y.YY   |    | $X+Y       (currency format)
    [bottom] Total               |   sum     |   sum     |    | grand sum  (bold)
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# rebate_master.finance_company -> (pretty bucket name, tab-name suffix)
# Keys are the stored data values (unchanged); the pretty names are display only.
BUCKET_PRETTY = {
    "Self-Financed": ("Self-Funded", "Self-Funded"),
    "NewLane Financed": ("Newlane Financed", "NewLane Fin"),
    "OnePlace Capital": ("OnePlace Capital", "OnePlace Ca"),
}

MONEY_FMT = '"$"#,##0.00'


def short_period(months: list[date]) -> str:
    """Tab-name period: 'Dec 2025 & Jan 2026'."""
    return " & ".join(m.strftime("%b %Y") for m in months)


def long_period(months: list[date]) -> str:
    """Title period: 'December 2025 and January 2026'."""
    parts = [m.strftime("%B %Y") for m in months]
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        return f"{parts[0]} and {parts[1]}"
    return ", ".join(parts[:-1]) + f", and {parts[-1]}"


def _month_label(d: date) -> str:
    """Per-column header: just the month name."""
    return d.strftime("%B")


def build(per_bucket: dict, months: list[date]) -> bytes:
    """Build the xlsx.

    per_bucket: {bucket_finance_company: {clinic_display_name: {month_label: amount, ...}}}
    months: ordered list of date(year, month, 1).
    """
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    total_font = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    box = Border(top=thin, bottom=thin, left=thin, right=thin)

    short = short_period(months)
    long_title = long_period(months)
    month_labels = [_month_label(m) for m in months]
    ncols = 1 + len(months) + 1  # Clinic + months + Amount

    # Stable bucket order to mirror the existing workbook
    bucket_order = ["OnePlace Capital", "NewLane Financed", "Self-Financed"]
    for bucket in bucket_order:
        if bucket not in per_bucket:
            continue
        pretty, suffix = BUCKET_PRETTY.get(bucket, (bucket, bucket[:18]))
        tab_name = f"{short} {suffix}"[:31]
        ws = wb.create_sheet(tab_name)

        # Row 2: merged title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        title_cell = ws.cell(row=2, column=1, value=f"{long_title} - {pretty} Rebates")
        title_cell.font = title_font
        title_cell.alignment = centered

        # Row 4: column headers
        ws.cell(row=4, column=1, value="Clinic/Hospital Name").font = header_font
        ws.cell(row=4, column=1).alignment = centered
        ws.cell(row=4, column=1).border = box
        for i, label in enumerate(month_labels, start=2):
            c = ws.cell(row=4, column=i, value=label)
            c.font = header_font; c.alignment = centered; c.border = box
        last_col = ncols
        c = ws.cell(row=4, column=last_col, value="Amount")
        c.font = header_font; c.alignment = centered; c.border = box

        # Row 6+: clinic rows
        clinics = per_bucket[bucket]
        sorted_clinics = sorted(clinics.items(), key=lambda kv: kv[0].lower())
        col_totals = {label: 0.0 for label in month_labels}
        grand_total = 0.0
        row = 6
        for clinic_name, month_amounts in sorted_clinics:
            ws.cell(row=row, column=1, value=clinic_name).border = box
            row_total = 0.0
            for i, label in enumerate(month_labels, start=2):
                v = round(float(month_amounts.get(label, 0.0)), 2)
                c = ws.cell(row=row, column=i, value=v)
                c.number_format = MONEY_FMT; c.border = box
                col_totals[label] += v
                row_total += v
            c = ws.cell(row=row, column=last_col, value=round(row_total, 2))
            c.number_format = MONEY_FMT; c.border = box
            grand_total += row_total
            row += 1

        # Total row
        c = ws.cell(row=row, column=1, value="Total")
        c.font = total_font; c.alignment = Alignment(horizontal="right"); c.border = box
        for i, label in enumerate(month_labels, start=2):
            c = ws.cell(row=row, column=i, value=round(col_totals[label], 2))
            c.number_format = MONEY_FMT; c.font = total_font; c.border = box
        c = ws.cell(row=row, column=last_col, value=round(grand_total, 2))
        c.number_format = MONEY_FMT; c.font = total_font; c.border = box

        # Column widths
        ws.column_dimensions["A"].width = 56
        for i in range(2, ncols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
